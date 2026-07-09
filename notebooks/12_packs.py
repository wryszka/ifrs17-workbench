# Databricks notebook source
# MAGIC %md
# MAGIC # 12 · Board pack — the Excel deliverable (connect to Excel, don't replace it)
# MAGIC
# MAGIC Builds the multi-sheet board pack from the engine tables: statements, §101 and §104
# MAGIC roll-forwards, RA roll-forward, recon, and a **provenance sheet** (the run-audit join —
# MAGIC every number's inputs pinned to Delta versions). Written to the `packs/` area of the
# MAGIC Volume as both .xlsx and the app download source. The sign-off CERTIFICATE PDF is built
# MAGIC by the app at sign-off time (server/packs.py) — this notebook is the recurring pack.

# COMMAND ----------

# MAGIC %pip install openpyxl --quiet

# COMMAND ----------

# MAGIC %run ./engine_common

# COMMAND ----------

import shutil
import tempfile

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

HDR = Font(bold=True, color="FFFFFF")
FILL = PatternFill("solid", fgColor="1E293B")
NUM = "#,##0.00"


def sheet_from_df(wb, title, df, money_cols):
    ws = wb.create_sheet(title)
    ws.append(list(df.columns))
    for c in range(1, len(df.columns) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font, cell.fill = HDR, FILL
    for _, r in df.iterrows():
        ws.append([r[c] for c in df.columns])
    for ci, col in enumerate(df.columns, start=1):
        width = max(14, min(52, int(df[col].astype(str).str.len().max() or 12) + 2))
        ws.column_dimensions[ws.cell(row=1, column=ci).column_letter].width = width
        if col in money_cols:
            for ri in range(2, len(df) + 2):
                ws.cell(row=ri, column=ci).number_format = NUM
    return ws


wb = Workbook()
ws0 = wb.active
ws0.title = "Cover"
ws0["A1"] = "Bricksurance SE — IFRS 17 board pack"
ws0["A1"].font = Font(bold=True, size=16)
ws0["A2"] = f"Close period {CLOSE_PERIOD} · reporting date {REPORTING_DATE} · produced by the IFRS 17 Workbench"
ws0["A4"] = ("About this pack: Bricksurance SE is a fictional insurer; every figure is computed by governed "
             "engines on synthetic data (real EIOPA discount curves). Illustrative configuration — not "
             "accounting advice.")
ws0["A4"].alignment = Alignment(wrap_text=True)
ws0.column_dimensions["A"].width = 110

pl = pdf(f"SELECT close_period, line_no, line_item, amount FROM {FQ}.gld_pnl_statement ORDER BY close_period, line_no")
sheet_from_df(wb, "P&L (§80)", pl, ["amount"])
bs = pdf(f"SELECT close_period, line_no, line_item, side, amount FROM {FQ}.gld_balance_sheet ORDER BY close_period, line_no")
sheet_from_df(wb, "Balance sheet", bs, ["amount"])
d1 = pdf(f"SELECT close_period, component, line, amount FROM {FQ}.gld_disclosure_lrc_lic ORDER BY close_period, component, ord")
sheet_from_df(wb, "Roll-forward (§101)", d1, ["amount"])
d4 = pdf(f"SELECT close_period, component, line, amount FROM {FQ}.gld_disclosure_by_component ORDER BY close_period, component, ord")
sheet_from_df(wb, "By component (§104)", d4, ["amount"])
ra = pdf(f"SELECT * FROM {FQ}.gld_ra_rollforward ORDER BY close_period")
sheet_from_df(wb, "Risk adjustment", ra, ["opening", "net_change", "closing"])
rec = pdf(f"SELECT close_period, recon_item, subledger_amount, gl_amount, raw_difference, journal_adjustment, residual, status "
          f"FROM {FQ}.gld_trial_balance_recon WHERE close_period = '{CLOSE_PERIOD}' ORDER BY recon_item")
sheet_from_df(wb, "GL reconciliation", rec,
              ["subledger_amount", "gl_amount", "raw_difference", "journal_adjustment", "residual"])
prov = pdf(f"""SELECT run_id, engine, close_period, status, input_versions, assumption_versions, curve_dates, note
               FROM {FQ}.gov_run_audit ORDER BY finished_at DESC LIMIT 40""")
ws = sheet_from_df(wb, "Provenance", prov, [])
ws["J1"] = "The audit trail is a join, not a project: engine run × pinned input Delta versions × approved assumption versions × curve dates."

with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tf:
    wb.save(tf.name)
    dest = f"{VOL}/packs/ifrs17_board_pack_{CLOSE_PERIOD}.xlsx"
    shutil.copyfile(tf.name, dest)
print("board pack →", dest)

set_status(10, "Group submission", "in_progress", "board pack produced; awaiting sign-off certificate", "packs")
log_run("board_pack",
        ["gld_pnl_statement", "gld_balance_sheet", "gld_disclosure_lrc_lic", "gld_disclosure_by_component",
         "gld_ra_rollforward", "gld_trial_balance_recon", "gov_run_audit"],
        {}, [], note=f"pack at {dest}")
print("12 complete")
