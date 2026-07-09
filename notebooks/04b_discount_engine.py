# Databricks notebook source
# MAGIC %md
# MAGIC # 04b · Discount + risk-adjustment engine
# MAGIC
# MAGIC 1. Parses the **real EIOPA monthly publications** on the Volume (`RFR_spot_no_VA`, EUR) →
# MAGIC    `ref_rfr_curve` + `ref_rfr_meta` (published UFR/LLP/CRA, shown on the Discount page).
# MAGIC 2. Builds `ref_discount_curve`: monthly discount factors per curve date × portfolio =
# MAGIC    EIOPA spot + **illiquidity premium bps** (versioned assumption). IFRS 17 rates are NOT
# MAGIC    Solvency II rates — the ILP overlay is the entity's own (illustrative) bottom-up choice.
# MAGIC 3. PVs the LRC-scope reserving projections per group × close quarter on BOTH bases
# MAGIC    (locked-in per cohort, current per close) + RA (CL-75 lognormal) → `gld_fcf_summary`
# MAGIC    and `gld_discount_impact` (the P&L/OCI split driver).
# MAGIC
# MAGIC Historical "current" bases use the nearest bundled real publication (disclosed).

# COMMAND ----------

# MAGIC %pip install openpyxl --quiet

# COMMAND ----------

# MAGIC %run ./engine_common

# COMMAND ----------

import glob
import os
import re

from openpyxl import load_workbook

rows_curve, rows_meta = [], []
for path in sorted(glob.glob(f"{VOL}/eiopa/EIOPA_RFR_*_Term_Structures.xlsx")):
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb["RFR_spot_no_VA"]
    eur_col = None
    for c in range(2, 40):
        if ws.cell(row=2, column=c).value == "Euro":
            eur_col = c
            break
    assert eur_col, f"No Euro column in {path}"
    code = ws.cell(row=3, column=eur_col).value  # e.g. EUR_30_06_2026_SWP_LLP_20_EXT_40_UFR_3.30
    m = re.search(r"(\d{2})_(\d{2})_(\d{4})", code)
    eff = f"{m.group(3)}-{m.group(2)}-{m.group(1)}"
    meta = {ws.cell(row=r, column=2).value: ws.cell(row=r, column=eur_col).value for r in range(4, 11)}
    rows_meta.append((eff, float(meta.get("UFR") or 0), int(meta.get("LLP") or 0),
                      int(meta.get("Convergence") or 0), float(meta.get("CRA") or 0),
                      os.path.basename(path)))
    r = 11
    while True:
        mat = ws.cell(row=r, column=2).value
        if not isinstance(mat, (int, float)) or mat > 60:
            break
        spot = ws.cell(row=r, column=eur_col).value
        if spot is not None:
            rows_curve.append((eff, "EUR", int(mat), round(float(spot), 6), os.path.basename(path)))
        r += 1
    wb.close()

assert len(rows_meta) == 5, f"expected the 5 bundled EIOPA publications, found {len(rows_meta)}"
spark.createDataFrame(rows_curve, "curve_date string, currency string, maturity_years int, spot_rate double, source_file string") \
    .write.mode("overwrite").option("overwriteSchema", "true").saveAsTable(f"{FQ}.ref_rfr_curve")
spark.createDataFrame(rows_meta, "curve_date string, ufr double, llp int, convergence int, cra_bps double, source_file string") \
    .write.mode("overwrite").option("overwriteSchema", "true").saveAsTable(f"{FQ}.ref_rfr_meta")
spark.sql(f"COMMENT ON TABLE {FQ}.ref_rfr_curve IS 'REAL EIOPA RFR_spot_no_VA EUR spot curves — unmodified publications, provenance in data/eiopa/PROVENANCE.md'")
print(f"ref_rfr_curve: {len(rows_curve)} points across {len(rows_meta)} real publications")

# COMMAND ----------

# MAGIC %md ## Portfolio discount curves — EIOPA base + ILP, monthly factors

# COMMAND ----------

asm = load_assumptions()
ILP = asm["illiquidity_premium"][1]

spots = {}
for r in rows_curve:
    spots.setdefault(r[0], {})[r[2]] = r[3]

curve_rows = []
for eff in sorted(spots):
    s = spots[eff]
    max_y = max(s)
    for port, ilp_bps in sorted({**ILP, "_BASE": 0}.items()):  # _BASE = pure EIOPA (SII crosswalk uses it)
        for m in range(1, 361):
            t = m / 12.0
            lo, hi = int(t), min(int(t) + 1, max_y)
            lo = max(lo, 1)
            s_lo, s_hi = s.get(lo, s[1]), s.get(hi, s[max_y])
            spot = s_lo if hi == lo else s_lo + (s_hi - s_lo) * (t - lo) / (hi - lo)
            if t < 1:
                spot = s[1]
            rate = spot + ilp_bps / 10000.0
            df = (1.0 + rate) ** (-t)
            curve_rows.append((eff, port, m, round(spot, 6), ilp_bps, round(df, 8)))

spark.createDataFrame(curve_rows,
                      "curve_date string, portfolio_id string, maturity_month int, base_spot double, "
                      "ilp_bps int, discount_factor double") \
    .write.mode("overwrite").option("overwriteSchema", "true").saveAsTable(f"{FQ}.ref_discount_curve")
spark.sql(f"COMMENT ON TABLE {FQ}.ref_discount_curve IS 'IFRS 17 discount curves: real EIOPA spot + illiquidity premium per portfolio (versioned assumption). Monthly discount factors to 30y.'")
print(f"ref_discount_curve: {len(curve_rows)} rows")

# COMMAND ----------

# MAGIC %md ## FCF summary per group × close × basis (LRC scope) + discount impact

# COMMAND ----------

curves = load_curves()
groups = pdf(f"SELECT * FROM {FQ}.gld_contract_groups WHERE measurement_model != 'LIC_ONLY'")
ra_cov = pdf(f"SELECT portfolio_id, cov FROM {FQ}.ref_ra_params").set_index("portfolio_id")["cov"]

proj = pdf(f"""
    SELECT run_id, CAST(as_of_date AS STRING) as_of_date, portfolio_id,
           cohort_or_accident_year cohort_year, CAST(projection_month AS STRING) projection_month,
           cf_type, SUM(amount) amount
    FROM {FQ}.slv_cashflow_projection WHERE scope = 'LRC'
    GROUP BY 1,2,3,4,5,6""")
proj["m"] = proj["projection_month"].map(lambda s: datetime.date.fromisoformat(s))

fcf_rows, di_rows = [], []
for _, g in groups.sort_values("group_id").iterrows():
    gid, port, cy = g["group_id"], g["portfolio_id"], int(g["cohort_year"])
    li_date = g["locked_in_curve_date"]
    cov = float(ra_cov[port])
    for ql_ in QL:
        run = f"RSV_{ql_}"
        asof = q_end(ql_)
        sub = proj[(proj["run_id"] == run) & (proj["portfolio_id"] == port) & (proj["cohort_year"] == cy)]
        if sub.empty and cy > int(ql_[:4]):
            continue  # cohort not yet recognised
        for basis, cd in (("locked_in", li_date), ("current", CURRENT_CURVE[ql_])):
            curve = curves[(cd, port)]
            pv_prem = pv([(r["m"], r["amount"]) for _, r in sub[sub["cf_type"] == "premium"].iterrows()], curve, asof)
            pv_clm = pv([(r["m"], r["amount"]) for _, r in sub[sub["cf_type"] == "claims"].iterrows()], curve, asof)
            pv_exp = pv([(r["m"], r["amount"]) for _, r in sub[sub["cf_type"] == "expense"].iterrows()], curve, asof)
            pv_out = round(pv_clm + pv_exp, 2)
            ra = round(pv_out * ra_factor(cov), 2)
            fcf = round(pv_out + ra - pv_prem, 2)
            fcf_rows.append(dict(group_id=gid, portfolio_id=port, close_period=ql_, basis=basis,
                                 curve_date=cd, pv_future_premiums=pv_prem, pv_future_claims=pv_clm,
                                 pv_future_expenses=pv_exp, risk_adjustment=ra, fcf_remaining=fcf))

fcf = pd.DataFrame(fcf_rows)
write_engine(fcf, "gld_fcf_summary",
             "group_id string, portfolio_id string, close_period string, basis string, curve_date string, "
             "pv_future_premiums double, pv_future_claims double, pv_future_expenses double, "
             "risk_adjustment double, fcf_remaining double",
             "Fulfilment cash flows for REMAINING COVERAGE per group × close × basis. FCF = PV(outflows) + RA "
             "- PV(inflows). Locked-in basis drives GMM P&L; current basis drives the BS + onerous test; the "
             "difference is the OCI disaggregation.")

# discount impact (GMM groups): current vs locked-in FCF difference = accumulated OCI driver
gmm = fcf[fcf["group_id"].isin(groups[groups["measurement_model"] == "GMM"]["group_id"])]
piv = gmm.pivot_table(index=["group_id", "portfolio_id", "close_period"], columns="basis",
                      values="fcf_remaining").reset_index()
piv["oci_balance"] = (piv["current"] - piv["locked_in"]).round(2)
piv = piv.sort_values(["group_id", "close_period"]).reset_index(drop=True)
piv["oci_in_period"] = piv.groupby("group_id")["oci_balance"].diff().fillna(piv["oci_balance"]).round(2)
di = piv.rename(columns={"locked_in": "fcf_locked_in", "current": "fcf_current"})
write_engine(di[["group_id", "portfolio_id", "close_period", "fcf_locked_in", "fcf_current",
                 "oci_balance", "oci_in_period"]],
             "gld_discount_impact",
             "group_id string, portfolio_id string, close_period string, fcf_locked_in double, "
             "fcf_current double, oci_balance double, oci_in_period double",
             "GMM OCI disaggregation: BS measures FCF at current rates, P&L accretes at locked-in — the "
             "difference accumulates in OCI. CSM is insensitive to current-rate moves (accretion is locked-in).")

log_run("discount_ra_engine",
        ["slv_cashflow_projection", "gov_assumption_registry", "ref_ra_params", "gld_contract_groups"],
        {"illiquidity_premium": 1, "risk_adjustment_cl": 1},
        ["ref_rfr_curve", "ref_rfr_meta", "ref_discount_curve", "gld_fcf_summary", "gld_discount_impact"],
        curve_dates={"locked_in": LOCKED_IN_CURVE, "current": CURRENT_CURVE[CLOSE_PERIOD]},
        note="real EIOPA publications parsed; ILP overlay v1; RA CL-75")
print("04b complete")
