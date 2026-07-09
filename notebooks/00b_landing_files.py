# Databricks notebook source
# MAGIC %md
# MAGIC # 00b · Landing files — scatter the truth into source-system extracts
# MAGIC
# MAGIC Renders `gen_truth_*` into the files a real IFRS 17 close ingests, on the Volume landing
# MAGIC area. **Sums tie back to the truth exactly** (largest-remainder allocation in cents), so
# MAGIC bottom-up engine results are byte-stable while every record looks — and is — individually
# MAGIC real. Sources produced:
# MAGIC
# MAGIC | Feed | Format | System flavour |
# MAGIC |---|---|---|
# MAGIC | Policy admin extracts | CSV per cohort year | policy admin core |
# MAGIC | Claims snapshot | JSON lines | Guidewire ClaimCenter-shaped |
# MAGIC | Claim transactions | CSV per year | claims finance feed |
# MAGIC | Actuarial projections | CSV per reserving run | reserving system export (LIC + LRC scopes) |
# MAGIC | GL trial balance | CSV per period | SAP-shaped |
# MAGIC | Reinsurance treaties | CSV | ceded-re register |
# MAGIC | FX rates | CSV | market data |
# MAGIC | Expense allocation | **XLSX workbook** + parsed CSV | finance planning |
# MAGIC | EIOPA RFR curves | XLSX (real publications) | regulator |
# MAGIC
# MAGIC Also **stages** (does not land) the schema-drifted Q2 2026 projection file used by the
# MAGIC "inject bad feed" demo lever.

# COMMAND ----------

# MAGIC %pip install openpyxl --quiet

# COMMAND ----------

dbutils.widgets.text("catalog", "lr_dev_aws_us_catalog")
dbutils.widgets.text("schema", "ifrs17_workbench")
dbutils.widgets.text("seed", "42")

catalog = dbutils.widgets.get("catalog")
schema = dbutils.widgets.get("schema")
SEED = int(dbutils.widgets.get("seed"))
FQ = f"{catalog}.{schema}"
VOL = f"/Volumes/{catalog}/{schema}/ifrs17_files"

import csv
import datetime
import glob
import io
import json
import os
import random
import shutil

rng = random.Random(SEED)

REPORTING_DATE = datetime.date(2026, 6, 30)
QUARTERS = [(2024, 1), (2024, 2), (2024, 3), (2024, 4), (2025, 1), (2025, 2), (2025, 3), (2025, 4), (2026, 1), (2026, 2)]
QL = [f"{y}Q{q}" for y, q in QUARTERS]

truth_cohort = {(r["portfolio_id"], r["cohort_year"]): r.asDict()
                for r in spark.table(f"{FQ}.gen_truth_cohort").collect()}
truth_acc = {(r["portfolio_id"], r["accident_quarter"]): r.asDict()
             for r in spark.table(f"{FQ}.gen_truth_accident").collect()}
truth_ro = [r.asDict() for r in spark.table(f"{FQ}.gen_truth_runoff").orderBy("accident_year").collect()]
pay_pattern = {}
for r in spark.table(f"{FQ}.gen_pay_pattern").collect():
    pay_pattern.setdefault(r["portfolio_id"], {})[r["dev_quarter"]] = r["paid_share"]
pay_pattern = {p: [v for _, v in sorted(d.items())] for p, d in pay_pattern.items()}
portfolios = {r["portfolio_id"]: r.asDict() for r in spark.table(f"{FQ}.ref_portfolio").collect()}
treaties = [r.asDict() for r in spark.table(f"{FQ}.gen_truth_treaties").collect()]

FLOOD_Q, FLOOD_ULT = "2026Q2", None
for (p, aq), r in truth_acc.items():
    if r["event_ultimate"] > 0:
        FLOOD_ULT = r["event_ultimate"]


def cents(x):
    return int(round(x * 100))


def scatter(total_cents, weights, floor_cents=0):
    """Largest-remainder: split total_cents across len(weights) parts ∝ weights, exact sum.
    floor_cents guarantees a minimum per part (taken from the largest parts) so no record
    rounds to ~zero and trips a DQ drop rule — the sum stays exact."""
    s = sum(weights)
    raw = [total_cents * w / s for w in weights]
    base = [int(r) for r in raw]
    rem = total_cents - sum(base)
    order = sorted(range(len(raw)), key=lambda i: (raw[i] - base[i]), reverse=True)
    for i in range(rem):
        base[order[i % len(base)]] += 1
    if floor_cents and len(base) * floor_cents < total_cents:
        big = sorted(range(len(base)), key=lambda i: base[i], reverse=True)
        for i in range(len(base)):
            if base[i] < floor_cents:
                need = floor_cents - base[i]
                for j in big:
                    give = min(need, base[j] - floor_cents)
                    if give > 0:
                        base[j] -= give
                        base[i] += give
                        need -= give
                    if need == 0:
                        break
    return base


def q_start(y, q):
    return datetime.date(y, 3 * q - 2, 1)


def q_end(y, q):
    nx = datetime.date(y + (1 if q == 4 else 0), 1 if q == 4 else 3 * q + 1, 1)
    return nx - datetime.timedelta(days=1)


def month_add(d, m):
    y, mo = d.year + (d.month - 1 + m) // 12, (d.month - 1 + m) % 12 + 1
    return datetime.date(y, mo, 1)


def put(path, text):
    with open(f"{VOL}/{path}", "w", encoding="utf-8") as f:
        f.write(text)
    print(f"  landed {path} ({len(text) // 1024} KB)")


def csv_text(header, rows):
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(header)
    w.writerows(rows)
    return buf.getvalue()

# COMMAND ----------

# MAGIC %md ## 1 · Policy admin extracts — one CSV per cohort year
# MAGIC Premium scattered per policy; sums tie exactly per cohort AND per writing quarter.
# MAGIC PAA lines: premium + acquisition cash at inception. CLT: 3 annual installments.
# MAGIC DEC: single premium at inception.

# COMMAND ----------

REGIONS = ["DE", "NL", "BE", "AT"]
REGION_W = [0.48, 0.26, 0.16, 0.10]
policies = {}  # portfolio -> list of dicts (needed for claims linkage)

for (port, cy), c in sorted(truth_cohort.items()):
    n, gwp_c = c["policy_count"], cents(c["gwp"])
    wp = [c["wq1"], c["wq2"], c["wq3"], c["wq4"]]
    counts = scatter(n, [w if w > 0 else 1e-9 for w in wp])
    q_gwp = scatter(gwp_c, [w if w > 0 else 1e-9 for w in wp])
    cov_m = portfolios[port]["coverage_months"]
    plist = []
    idx = 0
    for qi in range(4):
        if wp[qi] <= 0 or counts[qi] == 0:
            continue
        weights = [rng.lognormvariate(0, 0.55) for _ in range(counts[qi])]
        prem = scatter(q_gwp[qi], weights, floor_cents=12000)
        qs, qe = q_start(cy, qi + 1), q_end(cy, qi + 1)
        span = (qe - qs).days
        for k in range(counts[qi]):
            idx += 1
            pid = f"POL-{port}-{cy}-{idx:06d}"
            inc = qs + datetime.timedelta(days=rng.randint(0, span))
            premium = prem[k] / 100.0
            si_mult = {"MOT": 12, "PROP": 260, "LIAB": 420, "CLT": 600, "DEC": 900}[port]
            plist.append(dict(policy_id=pid, portfolio_id=port, cohort_year=cy,
                              inception_date=inc, expiry_date=month_add(inc, cov_m) - datetime.timedelta(days=1),
                              annual_premium=round(premium if port != "CLT" else premium / 3, 2),
                              total_premium=round(premium, 2),
                              installments=3 if port == "CLT" else 1,
                              sum_insured=round(premium * si_mult, -2),
                              acq_cost=round(premium * c["acq_ratio"], 2),
                              region=rng.choices(REGIONS, REGION_W)[0],
                              channel=rng.choices(["broker", "direct", "bank"], [0.55, 0.35, 0.10])[0]))
    policies.setdefault(port, []).extend(plist)

for cy in (2024, 2025, 2026):
    rows = []
    for port in sorted(policies):
        for p in policies[port]:
            if p["cohort_year"] == cy:
                rows.append([p["policy_id"], p["portfolio_id"], p["cohort_year"], p["inception_date"],
                             p["expiry_date"], p["total_premium"], p["annual_premium"], p["installments"],
                             p["sum_insured"], p["acq_cost"], p["region"], p["channel"]])
    put(f"landing/policy_admin/policy_admin_{cy}.csv",
        csv_text(["policy_id", "portfolio_id", "cohort_year", "inception_date", "expiry_date",
                  "total_premium", "annual_premium", "installments", "sum_insured", "acq_cost",
                  "region", "channel"], rows))
print({p: len(v) for p, v in policies.items()})

# COMMAND ----------

# MAGIC %md ## 2 · Claims (Guidewire-shaped JSONL snapshot) + claim transactions
# MAGIC Attritional per accident quarter + the June 2026 flood event + the run-off tail.
# MAGIC Per-claim payments follow the line's development pattern; only payments ≤ 2026-06-30
# MAGIC become transactions — the remainder is the future the reserving projections estimate.

# COMMAND ----------

AVG_SEV = {"MOT": 5500, "PROP": 4200, "LIAB": 11000, "CLT": 45000, "DEC": 95000, "RO": 60000}
CAUSES = {"MOT": ["collision", "theft", "windscreen", "vandalism"],
          "PROP": ["escape_of_water", "fire", "storm", "theft", "subsidence"],
          "LIAB": ["employers_liability", "public_liability", "product_liability"],
          "CLT": ["machinery_breakdown", "erection_all_risks", "professional_negligence"],
          "DEC": ["structural_defect", "waterproofing_failure", "foundation_defect"],
          "RO": ["employers_liability", "industrial_disease", "public_liability"]}

claims, transactions = [], []
cid = 0


def dev_quarters(ay, aq):
    """Quarters (label, index) from accident quarter forward."""
    out, y, q = [], ay, aq
    for i in range(60):
        out.append((f"{y}Q{q}", i))
        q += 1
        if q == 5:
            y, q = y + 1, 1
    return out


def add_claims(port, aql, ultimate, n_target, peril, cat_code, policy_pool, loss_year, loss_q):
    """Create claims summing exactly to `ultimate`, with per-claim dev payments."""
    global cid
    if ultimate <= 0 or n_target == 0:
        return
    weights = [rng.lognormvariate(0, 0.9) for _ in range(n_target)]
    sev = scatter(cents(ultimate), weights, floor_cents=15000)
    qs, qe = q_start(loss_year, loss_q), q_end(loss_year, loss_q)
    span = max(1, (qe - qs).days)
    pat = pay_pattern[port]
    for k in range(n_target):
        cid += 1
        ult_c = sev[k]
        loss_date = qs + datetime.timedelta(days=rng.randint(0, span))
        pol = rng.choice(policy_pool) if policy_pool else None
        # exact per-dev-quarter allocation of the claim ultimate
        alloc = scatter(ult_c, pat)
        paid_to_date_c, future_c = 0, 0
        for (lbl, di) in dev_quarters(loss_year, loss_q):
            if di >= len(alloc):
                break
            dy, dq = int(lbl[:4]), int(lbl[-1])
            pay_date = q_end(dy, dq) - datetime.timedelta(days=rng.randint(5, 60))
            if pay_date <= loss_date:
                pay_date = loss_date + datetime.timedelta(days=10)
            amt = alloc[di]
            if amt == 0:
                continue
            if pay_date <= REPORTING_DATE:
                paid_to_date_c += amt
                if pay_date >= datetime.date(2024, 1, 1):
                    transactions.append([f"TXN-{cid:07d}-{di}", f"cc:{700000 + cid}", pay_date,
                                         "claim_payment", round(amt / 100, 2), port])
            else:
                future_c += amt
        status = "closed" if future_c == 0 else "open"
        claims.append(dict(
            claim_id=f"cc:{700000 + cid}", portfolio_id=port, accident_quarter=aql,
            policy_id=(pol["policy_id"] if pol else f"POL-{port}-LEGACY-{cid:06d}"),
            loss_date=loss_date, reported_date=loss_date + datetime.timedelta(days=rng.randint(1, 21)),
            peril=peril or rng.choice(CAUSES[port]), catastrophe_code=cat_code,
            status=status, gross_ultimate=round(ult_c / 100, 2),
            paid_to_date=round(paid_to_date_c / 100, 2),
            case_reserve=round((ult_c - paid_to_date_c) / 100, 2),
            region=(pol["region"] if pol else "DE")))


for (port, aql), t in sorted(truth_acc.items()):
    ay, aq = int(aql[:4]), int(aql[-1])
    pool = [p for p in policies.get(port, [])
            if p["inception_date"] <= q_end(ay, aq) and p["expiry_date"] >= q_start(ay, aq)]
    n = max(3, int(round(t["attritional_ultimate"] / AVG_SEV[port])))
    add_claims(port, aql, t["attritional_ultimate"], n, None, None, pool, ay, aq)
    if t["event_ultimate"] > 0:
        n_ev = max(3, int(round(t["event_ultimate"] / 24000)))
        add_claims(port, aql, t["event_ultimate"], n_ev, "flood", "CAT-2026-FLOOD-CE", pool, ay, aq)

# Run-off: accident-era claims, opening paid share at engine start (2023-12-31), tail per RO pattern.
for r in truth_ro:
    ay = r["accident_year"]
    remaining0 = r["ultimate"] * (1 - r["paid_share_at_start"])
    n = max(5, int(round(r["ultimate"] / AVG_SEV["RO"] / 3)))
    weights = [rng.lognormvariate(0, 0.8) for _ in range(n)]
    ult = scatter(cents(r["ultimate"]), weights, floor_cents=50000)
    paid0 = scatter(cents(r["ultimate"] * r["paid_share_at_start"]), weights)
    tail = pay_pattern["RO"][-14:]  # remaining development, re-normalised below
    for k in range(n):
        cid += 1
        rem_c = ult[k] - paid0[k]
        alloc = scatter(rem_c, tail)
        paid_recent_c, future_c = 0, 0
        for di, amt in enumerate(alloc):
            qy, qq = QUARTERS[di] if di < len(QUARTERS) else (2026 + (di - 9) // 4, (di - 9) % 4 + 1)
            pay_date = q_end(qy, qq) - datetime.timedelta(days=rng.randint(5, 60))
            if amt == 0:
                continue
            if pay_date <= REPORTING_DATE:
                paid_recent_c += amt
                transactions.append([f"TXN-{cid:07d}-{di}", f"cc:{700000 + cid}", pay_date,
                                     "claim_payment", round(amt / 100, 2), "RO"])
            else:
                future_c += amt
        claims.append(dict(
            claim_id=f"cc:{700000 + cid}", portfolio_id="RO", accident_quarter=f"{ay}Q3",
            policy_id=f"POL-RO-{ay}-{k:05d}", loss_date=datetime.date(ay, 7, 15),
            reported_date=datetime.date(ay, 8, 1), peril=rng.choice(CAUSES["RO"]),
            catastrophe_code=None, status="closed" if future_c == 0 else "open",
            gross_ultimate=round(ult[k] / 100, 2),
            paid_to_date=round((paid0[k] + paid_recent_c) / 100, 2),
            case_reserve=round((ult[k] - paid0[k] - paid_recent_c) / 100, 2), region="DE"))

# land: one JSONL snapshot as at the reporting date + transactions CSV per calendar year
lines = []
for c in claims:
    lines.append(json.dumps({
        "publicID": c["claim_id"], "claimNumber": c["claim_id"].replace("cc:", "CLM-"),
        "policyNumber": c["policy_id"], "portfolio": c["portfolio_id"],
        "lossDate": str(c["loss_date"]), "reportedDate": str(c["reported_date"]),
        "accidentQuarter": c["accident_quarter"], "lossCause": c["peril"],
        "catastrophe": ({"code": c["catastrophe_code"], "name": "June 2026 Central European floods (synthetic)"}
                        if c["catastrophe_code"] else None),
        "state": c["status"], "grossPaidToDate": c["paid_to_date"], "caseReserve": c["case_reserve"],
        "region": c["region"], "_extractedAt": str(REPORTING_DATE)}))
put("landing/claims/claims_snapshot_20260630.jsonl", "\n".join(lines))

for yr in (2024, 2025, 2026):
    rows = [t for t in transactions if str(yr) == str(t[2])[:4]]
    rows.sort(key=lambda r: (str(r[2]), r[0]))
    put(f"landing/claim_transactions/claim_transactions_{yr}.csv",
        csv_text(["transaction_id", "claim_id", "payment_date", "transaction_type", "amount", "portfolio_id"], rows))
print(f"claims: {len(claims)}  transactions: {len(transactions)}")

# COMMAND ----------

# MAGIC %md ## 3 · Actuarial projections — one CSV per reserving run (LIC + LRC scopes)
# MAGIC The reserving-system export the workbench CONSUMES (it never computes reserves).
# MAGIC Ultimate estimates = truth × a deterministic maturity factor (young accident quarters
# MAGIC over-estimated slightly → favourable AvE run-off, like real books). The Q2 2026 run
# MAGIC carries assumption set v2: `flood_freq_property` (PROP forward LR 78%) and
# MAGIC `casualty_inflation_clt` (+8% future claims on CLT 2025) — the hero movements.

# COMMAND ----------

EST_FACTOR = {"short": [1.045, 1.015, 1.0], "long": [1.06, 1.04, 1.025, 1.012, 1.0]}


def est_factor(port, age_q):
    prof = portfolios[port]["settlement_profile"]
    f = EST_FACTOR["short" if prof == "short" else "long"]
    return f[min(age_q, len(f) - 1)]


def flood_lr(port, cy, run_label):
    """Forward loss ratio for LRC-scope projections, by assumption set at the run."""
    c = truth_cohort[(port, cy)]
    if run_label >= "2026Q2" and c["lr_v2"] != c["lr_v1"]:
        return c["lr_v2"]
    return c["lr_v1"]


for run_i, (ry, rq) in enumerate(QUARTERS):
    run_label = f"{ry}Q{rq}"
    run_id = f"RSV_{run_label}"
    asof = q_end(ry, rq)
    asm = ('{"flood_freq_property": 2, "casualty_inflation_clt": 2}' if run_label == "2026Q2"
           else '{"flood_freq_property": 1, "casualty_inflation_clt": 1}')
    rows = []

    # --- LIC scope: expected remaining payments per portfolio × accident year, monthly ---
    for (port, aql), t in sorted(truth_acc.items()):
        ay, aq = int(aql[:4]), int(aql[-1])
        if (ay, aq) > (ry, rq):
            continue
        age = (ry - ay) * 4 + (rq - aq)
        ult_true = t["attritional_ultimate"] + t["event_ultimate"]
        ult_est = round(ult_true * est_factor(port, age), 2)
        pat = pay_pattern[port]
        paid_share = sum(pat[:age + 1])
        rem_est = ult_est - ult_true * paid_share  # actual paid to date is truth
        if rem_est <= 100:
            continue
        rem_pat = pat[age + 1:] or [1.0]
        rem_alloc = scatter(cents(rem_est), [max(s, 1e-9) for s in rem_pat])
        for di, amt in enumerate(rem_alloc):
            if amt == 0:
                continue
            # spread the dev quarter across its 3 months
            for mi, mamt in enumerate(scatter(amt, [1, 1, 1])):
                if mamt:
                    rows.append([run_id, str(asof), "LIC", port, ay, aql,
                                 str(month_add(asof.replace(day=1), 1 + di * 3 + mi)),
                                 "claims", round(mamt / 100, 2), asm])
    # run-off LIC
    for r in truth_ro:
        ay = r["accident_year"]
        rem0 = cents(r["ultimate"] * (1 - r["paid_share_at_start"]))
        tail = pay_pattern["RO"][-14:]
        alloc = scatter(rem0, tail)
        future = alloc[run_i + 1:]
        for di, amt in enumerate(future):
            if amt == 0:
                continue
            for mi, mamt in enumerate(scatter(amt, [1, 1, 1])):
                if mamt:
                    rows.append([run_id, str(asof), "LIC", "RO", ay, f"{ay}Q3",
                                 str(month_add(asof.replace(day=1), 1 + di * 3 + mi)),
                                 "claims", round(mamt / 100, 2), asm])

    # --- LRC scope: remaining-coverage cashflows per cohort group, monthly ---
    for (port, cy), c in sorted(truth_cohort.items()):
        cov_m = portfolios[port]["coverage_months"]
        lr = flood_lr(port, cy, run_label)
        # future premiums within the boundary (CLT installments at anniversaries)
        for p_q in range(4):
            w = [c["wq1"], c["wq2"], c["wq3"], c["wq4"]][p_q]
            if w <= 0:
                continue
            # a reserving run can only project contracts that EXIST at its as-of date —
            # tranches not yet written are absent from earlier runs (they arrive as new business)
            if q_start(cy, p_q + 1) > asof:
                continue
            written = c["gwp"] * w
            inc_mid = q_start(cy, p_q + 1) + datetime.timedelta(days=45)
            if port == "CLT":
                for inst in (1, 2):
                    due = month_add(inc_mid.replace(day=1), 12 * inst)
                    if due > asof:
                        rows.append([run_id, str(asof), "LRC", port, cy, f"{port}-{cy}",
                                     str(due), "premium", round(written / 3, 2), asm])
            # expected future claims + expenses from UNEXPIRED coverage, monthly to expiry
            start_m = (cy - 2000) * 12 + p_q * 3 + 1
            end_m = start_m + cov_m
            asof_m = (asof.year - 2000) * 12 + asof.month
            rem_months = max(0, end_m - asof_m)
            if rem_months == 0:
                continue
            monthly_claims = written * lr / cov_m
            monthly_exp = written * c["expense_ratio"] / cov_m
            for m in range(1, rem_months + 1):
                rows.append([run_id, str(asof), "LRC", port, cy, f"{port}-{cy}",
                             str(month_add(asof.replace(day=1), m)), "claims",
                             round(monthly_claims, 2), asm])
                rows.append([run_id, str(asof), "LRC", port, cy, f"{port}-{cy}",
                             str(month_add(asof.replace(day=1), m)), "expense",
                             round(monthly_exp, 2), asm])

    rows.sort(key=lambda r: (r[2], r[3], str(r[4]), r[6], r[7]))
    put(f"landing/actuarial_projections/rsv_projection_{run_label}_v{'2' if run_label == '2026Q2' else '1'}.csv",
        csv_text(["run_id", "as_of_date", "scope", "portfolio_id", "cohort_or_accident_year", "group_ref",
                  "projection_month", "cf_type", "amount", "assumption_set"], rows))

# --- The DRIFTED Q2 2026 v1 file (staged, NOT landed): renamed columns + missing cf_type ---
drift_rows = []
asof = REPORTING_DATE
for i in range(200):
    drift_rows.append([f"RSV_2026Q2", str(asof), "LIC", "PROP", 2026, "2026Q2",
                       str(month_add(asof.replace(day=1), 1 + i % 12)), round(rng.uniform(5e4, 4e5), 2)])
with open(f"{VOL}/staging/rsv_projection_2026Q2_v1_DRIFTED.csv", "w") as f:
    w = csv.writer(f)
    w.writerow(["run_id", "as_of_date", "scope", "portfolio_id", "cohort_or_accident_year", "group_ref",
                "month", "cashflow_amt"])  # drift: projection_month→month, amount→cashflow_amt, cf_type MISSING
    w.writerows(drift_rows)
print("staged drifted file (not landed): staging/rsv_projection_2026Q2_v1_DRIFTED.csv")

# COMMAND ----------

# MAGIC %md ## 4 · GL trial balance, reinsurance register, FX, expense workbook
# MAGIC GL cash legs derive from the same truth (they tie by construction); the ONE deliberate
# MAGIC break: in Q2 2026 SAP books €412,340.00 of attributable claims-ops cost to other opex
# MAGIC (7000) — the recon finds it, an approved manual journal fixes it.

# COMMAND ----------

MISBOOKED = 412_340.00
COST_CENTRES = [("CC-100", "Claims operations", 0.34, True), ("CC-200", "Underwriting operations", 0.26, True),
                ("CC-300", "Policy administration", 0.20, True), ("CC-400", "IT & platforms", 0.12, True),
                ("CC-500", "Finance & actuarial", 0.08, True),
                ("CC-800", "Marketing & brand", None, False), ("CC-900", "Group overhead", None, False)]
ALLOC = {"MOT": 0.34, "PROP": 0.27, "LIAB": 0.12, "CLT": 0.14, "DEC": 0.08, "RO": 0.05}


def quarter_cash(y, q):
    """Truth-derived cash flows for the quarter (premiums received, claims paid, acq paid, expenses)."""
    prem = acq = 0.0
    for port, plist in policies.items():
        for p in plist:
            if q_start(y, q) <= p["inception_date"] <= q_end(y, q):
                prem += p["total_premium"] if p["installments"] == 1 else p["annual_premium"]
                acq += p["acq_cost"]
            elif p["installments"] > 1:  # CLT anniversaries
                for inst in (1, 2):
                    due = month_add(p["inception_date"].replace(day=1), 12 * inst)
                    if q_start(y, q) <= due <= q_end(y, q):
                        prem += p["annual_premium"]
    paid = sum(t[4] for t in transactions if q_start(y, q) <= t[2] <= q_end(y, q))
    earned_q = sum(t["earned_premium"] for (po, aq), t in truth_acc.items() if aq == f"{y}Q{q}")
    exp_att = sum(t["earned_premium"] * truth_cohort[(po, 2025)]["expense_ratio"]
                  for (po, aq), t in truth_acc.items() if aq == f"{y}Q{q}")
    return round(prem, 2), round(paid, 2), round(acq, 2), round(exp_att, 2), round(earned_q, 2)


gl_all, exp_feed = [], []
for (y, q) in QUARTERS:
    lbl = f"{y}Q{q}"
    prem, paid, acq, exp_att, earned = quarter_cash(y, q)
    non_att = round(1_900_000 + 130_000 * ((y - 2024) * 4 + q), 2)
    mis = MISBOOKED if lbl == "2026Q2" else 0.0
    ri_prem = round(sum(0.30 * truth_cohort[("PROP", cy)]["gwp"] / 4 for cy in (2025, 2026)
                        if not (cy == 2026 and y < 2026)) if y >= 2025 else 0.0, 2)
    gl_rows = [
        ("1000", "Cash and equivalents", round(prem - paid - acq - exp_att - non_att - ri_prem, 2)),
        ("4900", "Premiums received (statutory memo)", prem),
        ("5900", "Claims paid (statutory memo)", -paid),
        ("5040", "Attributable insurance service expenses", -(round(exp_att - mis, 2))),
        ("7000", "Other operating expenses", -(round(non_att + mis, 2))),
        ("5950", "Acquisition cash flows (memo)", -acq),
        ("2300", "Reinsurance premium ceded (memo)", -ri_prem),
    ]
    for acct, name, bal in gl_rows:
        gl_all.append([lbl, acct, name, "CC-000", bal])
    # expense allocation feed rows (per cost centre per quarter)
    att_total = exp_att
    for cc, name, share, is_att in COST_CENTRES:
        amt = round(att_total * share, 2) if is_att else (round(non_att * (0.55 if cc == "CC-800" else 0.45), 2))
        exp_feed.append([lbl, cc, name, amt, "attributable" if is_att else "non_attributable"])

for (y, q) in QUARTERS:
    lbl = f"{y}Q{q}"
    rows = [r for r in gl_all if r[0] == lbl]
    put(f"landing/gl_trial_balance/gl_trial_balance_{lbl}.csv",
        csv_text(["period", "gl_account", "gl_account_name", "cost_centre", "movement_eur"], rows))

put("landing/reinsurance/reinsurance_treaties.csv",
    csv_text(["treaty_id", "treaty_type", "portfolios", "cession_pct", "commission_pct",
              "inception", "expiry", "counterparty", "description"],
             [[t["treaty_id"], t["treaty_type"], t["portfolios"], t["cession_pct"], t["commission_pct"],
               t["inception"], t["expiry"], t["counterparty"], t["description"]] for t in treaties]))

fx = []
FX_BASE = {"GBP": 0.853, "USD": 1.084, "CHF": 0.938}
for i, (y, q) in enumerate(QUARTERS):
    for ccy, base in FX_BASE.items():
        fx.append([str(q_end(y, q)), f"EUR{ccy}", round(base + 0.006 * ((i * 7) % 5 - 2), 4)])
put("landing/fx_rates/fx_rates.csv", csv_text(["rate_date", "pair", "rate"], fx))

# Expense allocation WORKBOOK (xlsx) — the finance-planning artefact — plus a parsed CSV feed.
from openpyxl import Workbook

wb = Workbook()
ws1 = wb.active
ws1.title = "Allocation keys"
ws1.append(["cost_centre", "cost_centre_name", "attributable"] + list(ALLOC.keys()))
for cc, name, share, is_att in COST_CENTRES:
    ws1.append([cc, name, "yes" if is_att else "no"] + ([ALLOC[p] for p in ALLOC] if is_att else [0] * 6))
ws2 = wb.create_sheet("Quarterly amounts")
ws2.append(["period", "cost_centre", "cost_centre_name", "amount_eur", "classification"])
for r in exp_feed:
    ws2.append(r)
ws3 = wb.create_sheet("README")
ws3.append(["Bricksurance SE — expense allocation workbook (synthetic). Attributable cost centres allocate"])
ws3.append(["to portfolios by the keys on sheet 1; non-attributable flows to other operating expenses."])
import tempfile
with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tf:
    wb.save(tf.name)
    shutil.copyfile(tf.name, f"{VOL}/landing/expense_allocation/expense_allocation_2026Q2.xlsx")
print("  landed expense_allocation_2026Q2.xlsx")

alloc_rows = [[cc, name, "yes" if is_att else "no"] + ([ALLOC[p] for p in ALLOC] if is_att else [0] * 6)
              for cc, name, share, is_att in COST_CENTRES]
put("landing/expense_allocation/expense_allocation_keys.csv",
    csv_text(["cost_centre", "cost_centre_name", "attributable"] + list(ALLOC.keys()), alloc_rows))
put("landing/expense_allocation/expense_amounts.csv",
    csv_text(["period", "cost_centre", "cost_centre_name", "amount_eur", "classification"], exp_feed))

# COMMAND ----------

# MAGIC %md ## 5 · EIOPA curve files (real publications) → Volume

# COMMAND ----------

CANDIDATES = ["/Workspace/Shared/ifrs17_workbench/files/data/eiopa",
              os.path.abspath(os.path.join(os.getcwd(), "..", "data", "eiopa"))]
src_dir = next((d for d in CANDIDATES if os.path.isdir(d)), None)
assert src_dir, f"EIOPA files not found in {CANDIDATES}"
for f in sorted(glob.glob(f"{src_dir}/EIOPA_RFR_*.xlsx")):
    shutil.copyfile(f, f"{VOL}/eiopa/{os.path.basename(f)}")
print("eiopa volume:", sorted(os.listdir(f"{VOL}/eiopa")))

# COMMAND ----------

# MAGIC %md ## 6 · Seeded manual journal (the recon fix) — lands as a journal feed file

# COMMAND ----------

put("landing/gl_trial_balance/manual_journals_seed.csv",
    csv_text(["journal_id", "period", "gl_account_dr", "gl_account_cr", "amount_eur", "narrative",
              "posted_by", "approved_by", "status"],
             [["MJ-2026Q2-001", "2026Q2", "5040", "7000", MISBOOKED,
               "Reclass: claims-ops attributable cost misbooked to other opex (SAP CC-100 mapping fix)",
               "finance.control@bricksurance.example", "cfo.office@bricksurance.example", "approved"]]))

print("00b complete.")
